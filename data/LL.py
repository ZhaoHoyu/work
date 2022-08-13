# -*- coding: utf-8 -*-
"""
Created on Thu Jun 30 22:11:53 2022

@author: F1_001
"""

import numpy as np
import openpyxl as xl
import pandas as pd
import os,re,math,cv2

#保存路径设置
filesave2path = r"E:\test.xlsx"

#打开路径
filepath = r"E:\test"
data_list = os.listdir(filepath)


wb = xl.Workbook()
wb_sht = wb.active
wb_sht.title = 'LL'
wb_sht.cell(1, 15).value = 'CIE_RatioB'
wb_sht.cell(1, 16).value = 'CIE_RatioR'
wb_sht.cell(1, 17).value = 'R_target_Lv'
wb_sht.cell(1, 18).value = 'G_target_Lv'
wb_sht.cell(1, 19).value = 'B_target_Lv'
wb_sht.cell(1, 20).value = 'RGC'
wb_sht.cell(1, 21).value = 'GGC'
wb_sht.cell(1, 22).value = 'BGC'
wb_sht.cell(1, 23).value = 'Rduv'
wb_sht.cell(1, 24).value = 'Gduv'
wb_sht.cell(1, 25).value = 'Bduv'

raw1 = 1
raw2 = 1
col = 1
bd = 0
n = 0
for each_data in data_list:
    each_data_path = os.path.join(filepath, each_data)
    #print(each_data)
    if 'lvCIExy_' in each_data:
        flg = "lvCIExy_(.*)_202.*"
        name = re.compile(flg).findall(each_data)

        data = pd.read_csv(each_data_path,header=None)
        for k in range(14):
            for m in range(5):
                wb_sht.cell(m+n+1,k+1).value=data.iloc[[m],[k]].values[0][0]
        for i in range(1,5):
            W = float(data.iloc[[i],[2]].values[0][0])
            Wx = float(data.iloc[[i],[3]].values[0][0])
            Wy = float(data.iloc[[i],[4]].values[0][0])
            R  = float(data.iloc[[i],[5]].values[0][0])
            Rx= float(data.iloc[[i],[6]].values[0][0])
            Ry= float(data.iloc[[i],[7]].values[0][0])
            G= float(data.iloc[[i],[8]].values[0][0])
            Gx= float(data.iloc[[i],[9]].values[0][0])
            Gy= float(data.iloc[[i],[10]].values[0][0])
            B= float(data.iloc[[i],[11]].values[0][0])
            Bx= float(data.iloc[[i],[12]].values[0][0])
            By= float(data.iloc[[i],[13]].values[0][0])
            CIE_RatioB=By*(Wy*(Gx-Rx)+Ry*(Wx-Gx)+Gy*(Rx-Wx))/(Gy*(Wy*(Rx-Bx)+By*(Wx-Rx)+Ry*(Bx-Wx)))
            CIE_RatioR=CIE_RatioB*Ry*(Wx-Bx)/(By*(Rx-Wx))+Ry*(Wx-Gx)/(Gy*(Rx-Wx))
            CIE_RatioG=1
            count = CIE_RatioB+CIE_RatioG+CIE_RatioR
            R_target_Lv=W*CIE_RatioR/count
            G_target_Lv=W*CIE_RatioG/count
            B_target_Lv=W*CIE_RatioB/count
            RGC = R/R_target_Lv
            GGC = G/G_target_Lv
            BGC = B/B_target_Lv
            wb_sht.cell(i+1+n, 15).value = CIE_RatioB
            wb_sht.cell(i+1+n, 16).value = CIE_RatioR
            wb_sht.cell(i+1+n, 17).value = R_target_Lv
            wb_sht.cell(i+1+n, 18).value = G_target_Lv
            wb_sht.cell(i+1+n, 19).value = B_target_Lv
            wb_sht.cell(i+1+n, 20).value = RGC
            wb_sht.cell(i+1+n, 21).value = GGC
            wb_sht.cell(i+1+n, 22).value = BGC
            u0 = 4*Rx/(-2*Rx+12*Ry+3)
            v0 = 9*Ry/(-2*Rx+12*Ry+3)
            u1 = 4*Gx/(-2*Gx+12*Gy+3)
            v1 = 9*Gy/(-2*Gx+12*Gy+3)
            u2 = 4*Bx/(-2*Bx+12*By+3)
            v2 = 9*By/(-2*Bx+12*By+3)
            wb_sht.cell(i+1+n,30).value = u0
            wb_sht.cell(i+1+n,31).value = v0
            wb_sht.cell(i+1+n,32).value = u1
            wb_sht.cell(i+1+n,33).value = v1
            wb_sht.cell(i+1+n,34).value = u2
            wb_sht.cell(i+1+n,35).value = v2
        wb_sht.cell(1+n,1).value = name[0]
        wb_sht.cell(n+3,23).value=math.sqrt((float(wb_sht.cell(n+3,30).value)-float(wb_sht.cell(n+2,30).value))**2+(float(wb_sht.cell(n+3,31).value)-float(wb_sht.cell(n+2,31).value))**2)
        wb_sht.cell(n+4,23).value=math.sqrt((wb_sht.cell(n+4,30).value-wb_sht.cell(n+2,30).value)**2+(wb_sht.cell(n+4,31).value-wb_sht.cell(n+2,31).value)**2)
        wb_sht.cell(n+5,23).value=math.sqrt((wb_sht.cell(n+5,30).value-wb_sht.cell(n+2,30).value)**2+(wb_sht.cell(n+5,31).value-wb_sht.cell(n+2,31).value)**2)
        wb_sht.cell(n+3,24).value=math.sqrt((wb_sht.cell(n+3,32).value-wb_sht.cell(n+2,32).value)**2+(wb_sht.cell(n+3,33).value-wb_sht.cell(n+2,33).value)**2)
        wb_sht.cell(n+4,24).value=math.sqrt((wb_sht.cell(n+4,32).value-wb_sht.cell(n+2,32).value)**2+(wb_sht.cell(n+4,33).value-wb_sht.cell(n+2,33).value)**2)
        wb_sht.cell(n+5,24).value=math.sqrt((wb_sht.cell(n+4,32).value-wb_sht.cell(n+2,32).value)**2+(wb_sht.cell(n+5,33).value-wb_sht.cell(n+2,33).value)**2)
        wb_sht.cell(n+3,25).value=math.sqrt((wb_sht.cell(n+3,34).value-wb_sht.cell(n+2,34).value)**2+(wb_sht.cell(n+3,35).value-wb_sht.cell(n+2,35).value)**2)
        wb_sht.cell(n+4,25).value=math.sqrt((wb_sht.cell(n+4,34).value-wb_sht.cell(n+2,34).value)**2+(wb_sht.cell(n+4,35).value-wb_sht.cell(n+2,35).value)**2)
        wb_sht.cell(n+5,25).value=math.sqrt((wb_sht.cell(n+4,34).value-wb_sht.cell(n+2,34).value)**2+(wb_sht.cell(n+5,35).value-wb_sht.cell(n+2,35).value)**2)
        n += 6
    else:
        pass
wb.save(filesave2path)
print("Done")

