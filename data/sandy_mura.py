# -*- coding: utf-8 -*-
"""
Created on Fri Jul  1 12:16:06 2022

@author: F1_001
"""

import numpy as np
import openpyxl as xl
import pandas as pd
import statistics
import os,re,math,time

#保存路径设置
filesave2path = r"D:\data\LTPS_Tandem_15d5\puc\tet.xlsx"

#打开路径
filepath = r"D:\data\LTPS_Tandem_15d5\puc\ResultData"
data_list = os.listdir(filepath)


wb = xl.Workbook()
wb_sht = wb.active
wb_sht.title = 'sandy_mura'
wb_sht.cell(1, 2).value = 'min'
wb_sht.cell(1, 3).value = 'max'
wb_sht.cell(1, 4).value = 'avg'
wb_sht.cell(1, 5).value = 'std'
wb_sht.cell(1, 6).value = 'std/avg'

#2560*1600 max col=2460，max raw =1600
raw_start,raw_end = 600,1000
col_start,col_end = 1080,1480

wb_sht.cell(1, 8).value = str(str(raw_start)+'~'+str(raw_end)+','+str(col_start)+'~'+str(col_end))

time0 = time.time()
c=0
for each_data in data_list:
    each_data_path = os.path.join(filepath, each_data)
    #print(each_data)
    if 'PIEs_' in each_data:
        flg = "(.*).csv"
        name = re.compile(flg).findall(each_data)

        data = pd.read_csv(each_data_path,header=None)
        dt = []
        for i in range(raw_start,raw_end):
            for j in range(col_start,col_end):
                dt.append(data.iloc[[i],[j]].values[0][0])
        #min_Lv = min(dt)
        min_Lv = np.min(dt)
        max_Lv = np.max(dt)
        avg_Lv = np.mean(dt)
        #avg_Lv = sum(dt)/len(dt)
        std = statistics.pstdev(dt)
        std_avg = std/avg_Lv
        wb_sht.cell(c+2, 1).value = name[0]
        wb_sht.cell(c+2, 2).value = min_Lv
        wb_sht.cell(c+2, 3).value = max_Lv
        wb_sht.cell(c+2, 4).value = avg_Lv
        wb_sht.cell(c+2, 5).value = std
        wb_sht.cell(c+2, 6).value = std_avg
        
    c+=1
wb.save(filesave2path)
time1 = time.time()
t=time1-time0
print("total time = ",t)            