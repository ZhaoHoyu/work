# -*- coding: utf-8 -*-
"""
Created on Mon Jul 18 21:27:17 2022

@author: F1_001
"""


import numpy as np
import openpyxl as xl
import pandas as pd
import os,re,math

#保存路径设置
filesave2path = r"D:\CJM\LLC0719.xlsx"

#打开路径
filepath = r"D:\CJM\LLC0719"
data_list = os.listdir(filepath)


wb = xl.Workbook()
wb_sht = wb.active
wb_sht.title = 'uniformity'
wb_sht.cell(1, 3).value = 'dY'
wb_sht.cell(1, 4).value = 'duv'
wb_sht.cell(1, 5).value = 'min Lv'
wb_sht.cell(1, 6).value = 'max Lv'


raw1 = 1
raw2 = 1
raw3 = 1
raw4 = 1
col = 1
bd = 0
for each_data in data_list:
    each_data_path = os.path.join(filepath, each_data)
    #print(each_data)
    data = pd.read_csv(each_data_path,header=0)
    if col < 72:
        wb_sht.cell(1,6+col).value = ''
        wb_sht.cell(2,7+col).value = 'Lv'
        wb_sht.cell(2,8+col).value = 'x'
        wb_sht.cell(2,9+col).value = 'y'
    else:
        pass
    if "W_Band_11" in each_data:
        w = 0
        for i in range(10):
            for j in range(3):
                wb_sht.cell(raw1+2,8+j+w).value = float(data.iloc[[i],[3+j]].values[0][0])
            w += 4
        wb_sht.cell(1,8).value = "Band11"
        raw1 += 1
        col += 4
        bd += 15
    elif "W_Band_12" in each_data:
        w = 0
        for i in range(10):
            for j in range(3):
                wb_sht.cell(raw2+2,48+j+w).value = float(data.iloc[[i],[3+j]].values[0][0])
            w += 4
        wb_sht.cell(1,24).value = "Band12"
        raw2 += 1
        col += 4
        bd += 15
    elif "W_Band_15" in each_data:
        w = 0
        for i in range(3):
            for j in range(3):
                wb_sht.cell(raw3+2,88+j+w).value = float(data.iloc[[i],[3+j]].values[0][0])
            w += 4
        wb_sht.cell(1,40).value = "Band15"
        raw3 += 1
        col += 4
        bd += 15
    elif "W_Band_17" in each_data:
        w = 0
        for i in range(3):
            for j in range(3):
                wb_sht.cell(raw4+2,100+j+w).value = float(data.iloc[[i],[3+j]].values[0][0])
            w += 4
        wb_sht.cell(1,48).value = "Band17"
        raw4 += 1
        col += 4
        bd += 15
    else:
        pass


pt = filesave2path
wb.save(pt)
wb.close()

    
# = pd.readexcel(r'D:\Y\2_data\LTPS_15D5_Tandem\unif\LLC\LLC_26_band12.xlsx',header=None)
wb_ = xl.load_workbook(pt,read_only=False)
wb_cal = wb_.get_sheet_by_name("uniformity")
#sumar = wb_.create_sheet("sum")
#sumar.cell(1, 3).value = 'dY'
#sumar.cell(1, 4).value = 'duv'
#sumar.cell(1, 5).value = 'min Lv'
#sumar.cell(1, 6).value = 'max Lv'


wb_cal.cell(2, 1).value = 'Band11=1000Nits'
wb_cal.cell(2, 2).value = 'W8'
wb_cal.cell(3, 2).value = 'W10'
wb_cal.cell(4, 2).value = 'W51'
wb_cal.cell(5, 2).value = 'W68'
wb_cal.cell(6, 2).value = 'W86'
wb_cal.cell(7, 2).value = 'W87'
wb_cal.cell(8, 2).value = 'W172'
wb_cal.cell(9, 2).value = 'W192'
wb_cal.cell(10, 2).value = 'W215'
wb_cal.cell(11, 2).value = 'W255'
wb_cal.cell(13, 1).value = 'Band12=600Nits'
wb_cal.cell(13, 2).value = 'W8'
wb_cal.cell(14, 2).value = 'W10'
wb_cal.cell(15, 2).value = 'W51'
wb_cal.cell(16, 2).value = 'W68'
wb_cal.cell(17, 2).value = 'W86'
wb_cal.cell(18, 2).value = 'W87'
wb_cal.cell(19, 2).value = 'W172'
wb_cal.cell(20, 2).value = 'W192'
wb_cal.cell(21, 2).value = 'W215'
wb_cal.cell(22, 2).value = 'W255'
wb_cal.cell(24, 1).value = 'Band14=100Nits'
wb_cal.cell(24, 2).value = 'W10'
wb_cal.cell(25, 2).value = 'W87'
wb_cal.cell(26, 2).value = 'W255'
wb_cal.cell(28, 1).value = 'Band17=10Nits'
wb_cal.cell(28, 2).value = 'W15'
wb_cal.cell(29, 2).value = 'W23'
wb_cal.cell(30, 2).value = 'W255'



band_wt = 0
band_ht = 0
for b in range(4):
    g1 = 0
    c = 0
    #tbd=0
    if b==0:
        tbd=10
    elif b==1:
        tbd=10
    elif b==2:
        tbd=3
    elif b==3:
        tbd=3
    else:
        pass
    co = 0
    for g in range(tbd):
        Lv_list = []
        uv_list = []
        for p in range(15):
            x1 = wb_cal.cell(p+3,9+g1+band_wt).value
            y1 = wb_cal.cell(p+3,10+g1+band_wt).value
            u1 = 4*x1/(-2*x1+12*y1+3)
            v1 = 9*y1/(-2*x1+12*y1+3)
            Lv = wb_cal.cell(p+3,8+g1+band_wt).value
            Lv_list.append(Lv)
            for q in range(15):
                x2 = wb_cal.cell(q+3,9+g1+band_wt).value
                y2 = wb_cal.cell(q+3,10+g1+band_wt).value
                u2 = 4*x1/(-2*x2+12*y2+3)
                v2 = 9*y1/(-2*x2+12*y2+3)
                uv = math.sqrt((u2-u1)**2+(v2-v1)**2)
                uv_list.append(uv)
        min_Lv = min(Lv_list)
        max_Lv = max(Lv_list)
        dY = min_Lv/max_Lv
        duv = max(uv_list)
        print(dY,duv)
        wb_cal.cell(2+c+band_ht,3).value = dY
        wb_cal.cell(2+c+band_ht,4).value = duv
        wb_cal.cell(2+c+band_ht,5).value = min_Lv
        wb_cal.cell(2+c+band_ht,6).value = max_Lv
        g1 += 4
        c += 1
        co+=11
    tbd1=tbd*4
    tbd2=tbd+1
    band_wt += tbd1
    band_ht += tbd2
    

wb_.save(pt)
wb_.close()
